#!/usr/bin/env python3
"""Extract text content from various document formats.

Supported formats:
  - Text: txt, md, csv, json, yaml, yml, xml, log, and source code
  - Documents: doc, docx, rtf, odt, pages (via textutil)
  - Spreadsheets: xlsx (via openpyxl), xls (via textutil/strings)
  - Presentations: pptx (via zipfile XML), ppt (via textutil/strings)
  - PDF: via pdftotext

Usage: extract_text.py <file_path> [--max-chars N] [--sheet NAME]
"""

import sys
import os
import subprocess
import zipfile
import xml.etree.ElementTree as ET
import json
import re
import argparse
from pathlib import Path


def read_direct(filepath, max_chars):
    """Read text files directly."""
    encodings = ["utf-8", "gbk", "gb2312", "latin-1"]
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc) as f:
                return f.read(max_chars)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return f"[Error] Unable to decode {filepath} with supported encodings"


def read_csv(filepath, max_chars):
    """Read CSV with basic formatting."""
    import csv
    import io

    text = read_direct(filepath, max_chars)
    try:
        reader = csv.reader(io.StringIO(text))
        lines = []
        for row in reader:
            lines.append(" | ".join(row))
        return "\n".join(lines)
    except Exception:
        return text


def read_json(filepath, max_chars):
    """Read JSON with pretty formatting."""
    raw = read_direct(filepath, max_chars)
    try:
        data = json.loads(raw)
        return json.dumps(data, ensure_ascii=False, indent=2)[:max_chars]
    except json.JSONDecodeError:
        return raw


def extract_textutil(filepath, max_chars):
    """Extract text using macOS textutil (doc, docx, rtf, odt, pages)."""
    try:
        result = subprocess.run(
            ["textutil", "-convert", "txt", "-stdout", filepath],
            capture_output=True,
            text=True,
            timeout=30,
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout[:max_chars]
        return f"[Error] textutil failed: {result.stderr.strip()}"
    except FileNotFoundError:
        return "[Error] textutil not found"
    except subprocess.TimeoutExpired:
        return "[Error] textutil timed out"


def extract_pdf(filepath, max_chars):
    """Extract text from PDF using pdftotext."""
    try:
        result = subprocess.run(
            ["pdftotext", "-layout", filepath, "-"],
            capture_output=True,
            text=True,
            timeout=60,
        )
        if result.returncode == 0:
            text = result.stdout.strip()
            if text:
                return text[:max_chars]
            return "[Info] PDF contains no extractable text (may be scanned/image-based)"
        return f"[Error] pdftotext failed: {result.stderr.strip()}"
    except FileNotFoundError:
        # Fallback: try strings
        return _extract_strings(filepath, max_chars, "pdftotext not found, using strings fallback")
    except subprocess.TimeoutExpired:
        return "[Error] pdftotext timed out"


def extract_xlsx(filepath, max_chars):
    """Extract text from xlsx using openpyxl (if available) or XML parsing."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        output = []
        total_len = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output.append(f"=== Sheet: {sheet_name} ===")
            total_len += len(output[-1])

            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                # Skip completely empty rows
                if not any(cells):
                    continue
                line = " | ".join(cells)
                total_len += len(line) + 1
                if total_len > max_chars:
                    output.append("... [truncated]")
                    break
                output.append(line)

            output.append("")
            if total_len > max_chars:
                break

        wb.close()
        return "\n".join(output)

    except ImportError:
        return _extract_xlsx_zip(filepath, max_chars)
    except Exception as e:
        return f"[Error] Failed to read xlsx: {e}"


def _extract_xlsx_zip(filepath, max_chars):
    """Fallback xlsx extraction via zipfile XML parsing."""
    try:
        with zipfile.ZipFile(filepath, "r") as z:
            # Read shared strings
            strings = []
            if "xl/sharedStrings.xml" in z.namelist():
                tree = ET.parse(z.open("xl/sharedStrings.xml"))
                ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                for si in tree.findall(".//s:si", ns):
                    texts = si.findall(".//s:t", ns)
                    strings.append("".join(t.text or "" for t in texts))

            if strings:
                return "\n".join(strings)[:max_chars]
            return "[Info] No text content found in xlsx"
    except Exception as e:
        return f"[Error] Failed to parse xlsx: {e}"


def extract_xls(filepath, max_chars):
    """Extract text from old .xls format."""
    # Try textutil first
    text = extract_textutil(filepath, max_chars)
    if not text.startswith("[Error]"):
        return text
    return _extract_strings(filepath, max_chars, "xls format, using strings extraction")


def extract_pptx(filepath, max_chars):
    """Extract text from pptx via zipfile XML parsing."""
    try:
        with zipfile.ZipFile(filepath, "r") as z:
            slides = sorted(
                [f for f in z.namelist() if re.match(r"ppt/slides/slide\d+\.xml", f)]
            )

            output = []
            total_len = 0

            for slide_file in slides:
                slide_num = re.search(r"slide(\d+)", slide_file).group(1)
                output.append(f"=== Slide {slide_num} ===")

                tree = ET.parse(z.open(slide_file))
                # Extract all text from <a:t> elements
                texts = []
                for elem in tree.iter():
                    if elem.tag.endswith("}t") and elem.text:
                        texts.append(elem.text)

                slide_text = " ".join(texts)
                total_len += len(slide_text)
                if total_len > max_chars:
                    output.append(slide_text[: max_chars - total_len + len(slide_text)])
                    output.append("... [truncated]")
                    break
                output.append(slide_text)
                output.append("")

            return "\n".join(output) if output else "[Info] No text content found in pptx"
    except zipfile.BadZipFile:
        return "[Error] Invalid pptx file"
    except Exception as e:
        return f"[Error] Failed to parse pptx: {e}"


def extract_ppt(filepath, max_chars):
    """Extract text from old .ppt format."""
    text = extract_textutil(filepath, max_chars)
    if not text.startswith("[Error]"):
        return text
    return _extract_strings(filepath, max_chars, "ppt format, using strings extraction")


def _extract_strings(filepath, max_chars, note=""):
    """Last-resort extraction using strings command."""
    try:
        result = subprocess.run(
            ["strings", filepath],
            capture_output=True,
            text=True,
            timeout=30,
        )
        if result.stdout.strip():
            # Filter likely text (contains CJK or reasonable ASCII)
            lines = []
            for line in result.stdout.splitlines():
                line = line.strip()
                if len(line) >= 2 and not all(c in "._-/\\|{}[]<>()=" for c in line):
                    lines.append(line)
            text = "\n".join(lines)[:max_chars]
            prefix = f"[Note] {note}\n\n" if note else ""
            return prefix + text
        return "[Info] No extractable text found"
    except Exception as e:
        return f"[Error] strings extraction failed: {e}"


# Extension to extractor mapping
EXTRACTORS = {
    # Direct text
    ".txt": read_direct,
    ".md": read_direct,
    ".csv": read_csv,
    ".json": read_json,
    ".yaml": read_direct,
    ".yml": read_direct,
    ".xml": read_direct,
    ".log": read_direct,
    ".ini": read_direct,
    ".conf": read_direct,
    ".toml": read_direct,
    # Source code
    ".py": read_direct,
    ".swift": read_direct,
    ".js": read_direct,
    ".ts": read_direct,
    ".jsx": read_direct,
    ".tsx": read_direct,
    ".html": read_direct,
    ".css": read_direct,
    ".sh": read_direct,
    ".go": read_direct,
    ".rs": read_direct,
    ".java": read_direct,
    ".c": read_direct,
    ".h": read_direct,
    ".cpp": read_direct,
    ".m": read_direct,
    ".rb": read_direct,
    ".php": read_direct,
    ".sql": read_direct,
    # Documents (via textutil)
    ".doc": extract_textutil,
    ".docx": extract_textutil,
    ".rtf": extract_textutil,
    ".odt": extract_textutil,
    ".pages": extract_textutil,
    # Spreadsheets
    ".xlsx": extract_xlsx,
    ".xls": extract_xls,
    # Presentations
    ".pptx": extract_pptx,
    ".ppt": extract_ppt,
    # PDF
    ".pdf": extract_pdf,
}


def extract(filepath, max_chars=50000):
    """Extract text from file based on extension."""
    ext = Path(filepath).suffix.lower()
    extractor = EXTRACTORS.get(ext)

    if extractor is None:
        # Try direct read for unknown text-like files, otherwise strings
        try:
            return read_direct(filepath, max_chars)
        except Exception:
            return _extract_strings(filepath, max_chars, f"Unknown format {ext}")

    return extractor(filepath, max_chars)


def main():
    parser = argparse.ArgumentParser(description="Extract text from documents")
    parser.add_argument("filepath", help="Path to the file")
    parser.add_argument(
        "--max-chars",
        type=int,
        default=50000,
        help="Maximum characters to extract (default: 50000)",
    )
    args = parser.parse_args()

    filepath = os.path.expanduser(args.filepath)

    if not os.path.exists(filepath):
        print(f"[Error] File not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    text = extract(filepath, args.max_chars)

    # Output header
    size = os.path.getsize(filepath)
    size_str = (
        f"{size / 1048576:.1f} MB"
        if size >= 1048576
        else f"{size / 1024:.1f} KB" if size >= 1024 else f"{size} B"
    )
    ext = Path(filepath).suffix.lower()
    print(f"File: {os.path.basename(filepath)}")
    print(f"Path: {filepath}")
    print(f"Format: {ext} | Size: {size_str}")
    print(f"{'=' * 60}")
    print(text)


if __name__ == "__main__":
    main()
